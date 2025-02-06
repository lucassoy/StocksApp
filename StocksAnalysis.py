import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from plyer import notification
import yfinance as yf

# Función para obtener el valor del VIX usando yfinance
def obtener_valor_vix():
    try:
        vix = yf.Ticker("^VIX")
        vix_value = vix.history(period="1d")["Close"].iloc[-1]
        return float(vix_value)
    except Exception as e:
        print(f"Error al obtener el valor del VIX: {e}")
        return None

# Verificar el valor del VIX y enviar notificación si supera 25
vix_value = obtener_valor_vix()
if vix_value and vix_value > 25:
    notification.notify(
        title="Alerta de VIX",
        message=f"El índice VIX ha superado 25. Valor actual: {vix_value}",
        timeout=10
    )

# Cargar datos desde el Excel generado previamente
df = pd.read_excel("acciones.xlsx")

# Asegurar que las columnas necesarias existen y limpiar espacios
df["Symbol"] = df["Symbol"].astype(str).str.strip()
df["BPS1"] = pd.to_numeric(df["BPS1"], errors='coerce')
df["BPS2"] = pd.to_numeric(df["BPS2"], errors='coerce')
df["BPS3"] = pd.to_numeric(df["BPS3"], errors='coerce')
df["BPS4"] = pd.to_numeric(df["BPS4"], errors='coerce')
df["P/E Actual"] = pd.to_numeric(df["P/E Actual"], errors='coerce')
df["P/E Histórico"] = pd.to_numeric(df["P/E Histórico"], errors='coerce')

# Verificar los datos cargados
print("Datos cargados:")
print(df.head())

# Aplicar condiciones de inversión
df_recomendadas = df[
    (df['BPS1'] / df['BPS4'] > 1.5) &  # Crecimiento de 1.5x en 4 años
    (df['BPS1'] > df['BPS2']) &
    (df['BPS2'] > df['BPS3']) &
    (df['BPS3'] > df['BPS4'])  # Tendencia creciente del BPS
]

# Verificar las condiciones aplicadas
print("Condiciones aplicadas:")
print(df_recomendadas.head())

# Guardar las acciones recomendadas en un nuevo archivo Excel
df_recomendadas.to_excel("recomendaciones.xlsx", index=False)

# Cargar el archivo de recomendaciones para aplicar formato
wb = load_workbook("recomendaciones.xlsx")
ws = wb.active

# Listas para notificaciones
acciones_comprar = []
acciones_vender = []

# Aplicar formato y determinar acciones a comprar
for idx, row in df_recomendadas.iterrows():
    pe_actual = row["P/E Actual"]
    pe_historico = row["P/E Histórico"]
    if pd.notna(pe_actual) and pd.notna(pe_historico) and (pe_actual * 2 < pe_historico):
        # Marcar la fila en el Excel
        for cell in ws[idx + 2]:  # +2 para saltar el encabezado y ajustar el índice
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        acciones_comprar.append(row['Symbol'])

# Guardar el archivo de recomendaciones actualizado
wb.save("recomendaciones.xlsx")

# Cargar el archivo de compras para determinar acciones a vender
df_compras = pd.read_excel("compras.xlsx")

# Determinar acciones a vender
for idx, row in df_compras.iterrows():
    pe_actual = df.loc[df["Symbol"] == row["Symbol"], "P/E Actual"].values[0]
    pe_historico = df.loc[df["Symbol"] == row["Symbol"], "P/E Histórico"].values[0]
    if pd.notna(pe_actual) and pd.notna(pe_historico) and (pe_actual * 1.5 > pe_historico):
        acciones_vender.append(row['Symbol'])

# Enviar notificación de acciones a comprar
if acciones_comprar:
    notification.notify(
        title="Acciones recomendadas para compra",
        message=f"Las siguientes acciones deberían comprarse: {', '.join(acciones_comprar)}",
        timeout=10
    )

# Enviar notificación de acciones a vender
if acciones_vender:
    notification.notify(
        title="Acciones recomendadas para venta",
        message=f"Las siguientes acciones deberían venderse: {', '.join(acciones_vender)}",
        timeout=10
    )

# Mostrar las acciones recomendadas
print("📊 Acciones recomendadas para inversión:")
print(df_recomendadas[['Symbol', 'Name', 'Price', 'BPS1', 'BPS2', 'BPS3', 'BPS4', 'P/E Actual', 'P/E Histórico']])
