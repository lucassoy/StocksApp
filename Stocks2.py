import yfinance as yf
from openpyxl import Workbook, load_workbook
import time
import random
from bs4 import BeautifulSoup
import requests

# URL de Wikipedia con la lista de acciones del S&P 500
URL_SP500 = 'https://en.wikipedia.org/wiki/List_of_S%26P_500_companies'

# Obtener la tabla de Wikipedia
response = requests.get(URL_SP500)
soup = BeautifulSoup(response.text, 'html.parser')
tabla = soup.find('table')

# Verificar si la tabla fue encontrada
if not tabla:
    raise Exception("No se encontró la tabla en la página de Wikipedia.")

filas = tabla.find_all('tr')

# Extraer solo los símbolos y nombres de las acciones
datos = [["Symbol", "Name"]]
for fila in filas[1:]:  # Saltamos la primera fila de encabezados
    celdas = fila.find_all('td')
    if len(celdas) > 1:
        symbol = celdas[0].text.strip()
        name = celdas[1].text.strip()
        datos.append([symbol, name])

# Guardar en un archivo Excel
wb = Workbook()
ws = wb.active
for row in datos:
    ws.append(row)

# Agregar encabezados para los nuevos datos
ws.cell(row=1, column=3, value='Price')  # Precio actual
ws.cell(row=1, column=4, value='BPS1')
ws.cell(row=1, column=5, value='BPS2')
ws.cell(row=1, column=6, value='BPS3')
ws.cell(row=1, column=7, value='BPS4')
ws.cell(row=1, column=8, value='P/E Actual')
ws.cell(row=1, column=9, value='P/E Histórico')

wb.save('acciones.xlsx')

# Cargar archivo Excel
wb = load_workbook('acciones.xlsx')
ws = wb.active

# Iterar sobre los símbolos de las acciones
for idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=2):
    symbol = row[0]
    print(f'🔍 Buscando datos para {symbol}...')

    try:
        stock = yf.Ticker(symbol)

        # Obtener el precio actual
        price = stock.history(period="1d")['Close'].iloc[-1]
        ws.cell(row=idx, column=3, value=price)
        print(f'✅ {symbol} - Precio: {price}')

        # Obtener EPS de los últimos 5 años
        earnings = stock.financials.loc['Diluted EPS'].dropna().tolist()[:5]  # Tomar los últimos 5 años
        for i, eps in enumerate(earnings):
            ws.cell(row=idx, column=4 + i, value=eps)
        print(f'✅ {symbol} - EPS Históricos: {earnings}')

        # Obtener P/E Actual
        pe_actual = stock.info.get("trailingPE", "N/A")
        ws.cell(row=idx, column=8, value=pe_actual)
        print(f'✅ {symbol} - P/E Actual: {pe_actual}')

        # Calcular P/E Histórico
        pe_historico = []
        for i, eps in enumerate(earnings):
            if eps > 0:
                pe_historico.append(price / eps)
        if pe_historico:
            pe_historico_avg = sum(pe_historico) / len(pe_historico)
        else:
            pe_historico_avg = "N/A"
        ws.cell(row=idx, column=9, value=pe_historico_avg)
        print(f'✅ {symbol} - P/E Histórico: {pe_historico_avg}')

    except Exception as e:
        print(f'❌ Error al obtener datos de {symbol}: {e}')

# Guardar los datos en el Excel
wb.save('acciones.xlsx')
print('\n✅ Proceso completado. Datos guardados en "acciones.xlsx".')
